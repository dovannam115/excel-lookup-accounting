import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Excel Lookup Tool", layout="centered")

page_bg_img = '''
<style>
body {
    background-image: url("https://novajob.vn/uploads/recruitment/ntd/hoang_gia_luat_logo-1.png");
    background-size: contain;
    background-position: center top;
    background-repeat: no-repeat;
    background-attachment: fixed;
    background-color: #f5f5f5;
}
</style>
'''
st.markdown(page_bg_img, unsafe_allow_html=True)

st.title("🔍 Excel Lookup Tool")

option = st.radio("📌 Chọn chức năng", ["🔁 Lookup Bán ra & NXT", "📄 Lookup Mua vào & NXT"])

# --- Chức năng 1 ---
if option == "🔁 Lookup Bán ra & NXT":
    ban_ra_file = st.file_uploader("📤 Upload file Bán ra", type=["xlsx"], key="ban_ra")
    nxt_t4_file = st.file_uploader("📤 Upload file NXT", type=["xlsx"], key="nxt_t4")

    if ban_ra_file and nxt_t4_file:
        if st.button("🚀 Chạy Lookup"):
            ban_ra_df = pd.read_excel(ban_ra_file, sheet_name="Smart_KTSC_OK")
            nxt_t4_df = pd.read_excel(nxt_t4_file, sheet_name="F8_D", skiprows=22)
            nxt_t4_df.columns.values[[2, 4, 14]] = ['target_col', 'match_col', 'compare_col']

            q_col = ban_ra_df.columns[16]
            z_col = ban_ra_df.columns[25]

            results = []
            for _, row in ban_ra_df.iterrows():
                q_value = row[q_col]
                z_value = row[z_col]
                mask = (nxt_t4_df['match_col'] == q_value) & (nxt_t4_df['compare_col'] <= z_value)
                filtered = nxt_t4_df[mask].copy()
                if not filtered.empty:
                    filtered['diff'] = z_value - filtered['compare_col']
                    matched_row = filtered.loc[filtered['diff'].idxmin()]
                    results.append(matched_row['target_col'])
                else:
                    results.append("Không tìm thấy")
            
            ban_ra_df['lookup_result'] = results

            ban_ra_file.seek(0)
            wb = load_workbook(filename=ban_ra_file)
            if "Smart_KTSC_OK" in wb.sheetnames:
                ws = wb["Smart_KTSC_OK"]
                wb.remove(ws)
            ws_new = wb.create_sheet("Smart_KTSC_OK")
            for r in dataframe_to_rows(ban_ra_df, index=False, header=True):
                ws_new.append(r)
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("✅ DONE")
            st.download_button(
                label="📥 Tải file kết quả",
                data=output,
                file_name="BAN_RA_lookup_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# --- Chức năng 2 ---
elif option == "📄 Lookup Mua vào & NXT":
    data_file = st.file_uploader("📤 Upload file Data", type=["xlsx"], key="data")
    mapping_file = st.file_uploader("📤 Upload file Mapping", type=["xlsx"], key="mapping")

    error_threshold = st.number_input("🔧 Nhập phần trăm sai số cho phép (vd: 3% là 0.03)", min_value=0.0, max_value=1.0, value=0.03, step=0.01)

    if data_file and mapping_file:
        if st.button("🚀 Chạy Lookup Mapping"):
            data_df = pd.read_excel(data_file)
            mapping_df = pd.read_excel(mapping_file)

            # Đặt tên cột đúng theo yêu cầu
            data_df.columns.values[[16, 25]] = ['TENDM', 'DGVND']
            mapping_df.columns.values[[2, 4, 14]] = ['target_col', 'match_col', 'compare_col']

            # Hàm mô phỏng chính xác công thức Excel MATCH(1,...)
            def clean_text(val):
                if isinstance(val, str):
                    return val.strip().replace("\xa0", "").replace("\n", "").replace("\r", "")
                return val
            
            def lookup(row):
                try:
                    tendm = clean_text(row['TENDM'])
                    dgvnd = row['DGVND']
            
                    if dgvnd == 0 or pd.isna(dgvnd):
                        return "Không tìm thấy"
            
                    mapping_df_clean = mapping_df.copy()
                    mapping_df_clean['match_col'] = mapping_df_clean['match_col'].apply(clean_text)
                    mapping_df_clean['compare_col'] = pd.to_numeric(mapping_df_clean['compare_col'], errors='coerce')
            
                    filtered = mapping_df_clean[
                        (mapping_df_clean['match_col'] == tendm) &
                        (mapping_df_clean['compare_col'].notnull()) &
                        (abs(mapping_df_clean['compare_col'] - dgvnd) / dgvnd <= error_threshold)
                    ]
                    
                    if not filtered.empty:
                        return filtered.iloc[0]['target_col']
                    else:
                        return "Không tìm thấy"
                except:
                    return "Không tìm thấy"

            data_df['lookup_result'] = data_df.apply(lookup, axis=1)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                data_df.to_excel(writer, index=False, sheet_name="Data_Result")
            output.seek(0)

            st.success("✅ Lookup thành công!")
            st.download_button(
                label="📥 Tải file kết quả",
                data=output,
                file_name="data_lookup_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
